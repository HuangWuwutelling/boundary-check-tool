# 文档说明
# 请选择边界文件zip和地块信息.xlsx的文件地址，每个zip的文件名称应该为*X.zip，*为初步调查/详细调查等，X为13位数的地块编码。
# 地块信息.xlsx文件地址与边界文件zip相同，地块信息.xlsx应至少含地块编码、经度、纬度列，通过地块编码匹配zip文件，一行对应一个zip文件。
# 提示用户选择zip和地块信息.xlsx保存的文件夹路径
import os
import re
import sys
import zipfile

# 尝试导入模块，如果失败则给出友好提示
try:
    import geopandas as gpd
    import pandas as pd
    import folium
    import tkinter as tk
    import shapefile
    from tkinter import ttk, filedialog
    from shapely.geometry import Point
    import shutil
    import glob
    from collections import defaultdict
    import time
    import traceback
    import chardet  # 用于自动检测编码
    from typing import Optional
except ImportError as e:
    print(f"缺少必要的依赖库: {e}")
    print("请确保已安装所有依赖后再运行程序")
    print("如果是EXE版本，请联系作者获取完整版本")
    input("按任意键退出...")
    sys.exit(1)

TEMP_DIR = "temp_zip_extract"


def select_input_method() -> Optional[str]:
    """弹窗询问用户选择输入方式"""
    root = tk.Tk()
    root.title("选择输入方式")
    root.geometry("400x200")

    result = [None]  # 使用列表存储结果
    method = tk.StringVar(value="select")

    def on_confirm():
        if method.get() == "input":
            input_window = tk.Toplevel(root)
            input_window.title("输入文件夹路径")
            input_window.geometry("400x200")

            tk.Label(input_window, text="请输入文件夹路径:").pack(pady=5)
            path_entry = tk.Entry(input_window, width=50)
            path_entry.pack(pady=5, padx=10, fill="x")

            def on_input_confirm():
                # 修复：使用 result[0] 而不是 result["path"]
                result[0] = path_entry.get().strip()
                input_window.destroy()
                root.destroy()

            tk.Button(input_window, text="确认", command=on_input_confirm).pack(pady=5)

        else:  # "select"
            # 修复：使用 result[0] 而不是 result["path"]
            result[0] = filedialog.askdirectory(title="选择边界文件zip和地块信息.xlsx的文件夹", initialdir=".")
            root.destroy()

    tk.Label(root, text="请选择输入方式:", pady=10).pack()

    tk.Radiobutton(root, text="选择文件位置", variable=method, value="select").pack(anchor="w", padx=30)
    tk.Radiobutton(root, text="输入文件地址", variable=method, value="input").pack(anchor="w", padx=30)

    tk.Button(root, text="确认", command=on_confirm).pack(pady=10)

    root.wait_window()  # 等待窗口关闭
    return result[0]  # 返回结果


def select_folder_path():
    """选择包含zip文件和地块信息.xlsx的文件夹（支持两种方式）"""
    folder_path = select_input_method()

    if not folder_path:
        print("用户取消了操作或未选择文件夹，程序将结束。")
        sys.exit()
    print(f"用户选择的文件夹路径是: {folder_path}")
    return folder_path


def safe_remove_temp_dir():
    """安全删除临时目录，处理Windows文件锁定问题"""
    if not os.path.exists(TEMP_DIR):
        return

    max_attempts = 5
    for attempt in range(max_attempts):
        try:
            shutil.rmtree(TEMP_DIR)
            return  # 成功删除，退出函数
        except Exception as e:
            if attempt < max_attempts - 1:
                print(f"尝试删除临时目录失败 (第{attempt + 1}次): {str(e)}")
                time.sleep(1)  # 等待1秒再试
            else:
                print(f"无法删除临时目录 {TEMP_DIR}，请手动删除")
                # 尝试仅清空目录内容
                try:
                    for root, dirs, files in os.walk(TEMP_DIR):
                        for file in files:
                            file_path = os.path.join(root, file)
                            try:
                                os.chmod(file_path, 0o777)
                                os.remove(file_path)
                            except Exception as e:
                                print(f"无法删除文件 {file_path}: {str(e)}")
                    # 再次尝试删除目录
                    shutil.rmtree(TEMP_DIR)
                except:
                    pass


def detect_encoding(file_path):
    """检测文件编码"""
    try:
        # 尝试从.cpg文件获取编码
        cpg_path = os.path.splitext(file_path)[0] + ".cpg"
        if os.path.exists(cpg_path):
            try:
                with open(cpg_path, 'r', encoding='utf-8', errors='ignore') as f:
                    encoding = f.read().strip().lower()
                    if encoding in ['gbk', 'gb2312', 'cp936', 'ansi']:
                        return 'gbk'
                    elif encoding in ['utf-8', 'utf8']:
                        return 'utf-8'
                    elif encoding in ['latin1', 'iso-8859-1']:
                        return 'latin1'
            except:
                pass

        # 尝试自动检测DBF文件编码
        dbf_path = os.path.splitext(file_path)[0] + ".dbf"
        if os.path.exists(dbf_path):
            try:
                with open(dbf_path, 'rb') as f:
                    raw_data = f.read(10000)  # 读取前10000字节
                    result = chardet.detect(raw_data)
                    encoding = result['encoding'].lower() if result['encoding'] else 'utf-8'
                    confidence = result['confidence']

                    # 如果置信度高，直接使用
                    if confidence > 0.7:
                        if 'gb' in encoding or 'cp936' in encoding:
                            return 'gbk'
                        return encoding

                    # 如果置信度低，尝试常见编码
                    for enc in ['gbk', 'utf-8', 'cp936', 'latin1', 'mbcs']:
                        try:
                            with open(dbf_path, 'r', encoding=enc) as test:
                                test.read(100)
                            return enc
                        except:
                            continue
            except:
                pass

        # 默认返回GBK（中国常用）
        return 'gbk'
    except:
        return 'gbk'  # 默认使用GBK


def convert_timestamps_to_strings(gdf):
    """将GeoDataFrame中的所有Timestamp对象转换为字符串"""
    for col in gdf.columns:
        if col == 'geometry':
            continue

        # 检查是否是datetime类型
        if pd.api.types.is_datetime64_any_dtype(gdf[col]):
            try:
                # 尝试转换为YYYY-MM-DD格式
                gdf[col] = gdf[col].dt.strftime('%Y-%m-%d')
            except:
                # 如果转换失败，尝试直接转换为字符串
                gdf[col] = gdf[col].astype(str)
        # 检查是否是pandas Timestamp
        elif gdf[col].dtype == 'object':
            # 检查前几行是否是Timestamp
            sample = gdf[col].dropna().head(1)
            if not sample.empty and isinstance(sample.iloc[0], pd.Timestamp):
                try:
                    gdf[col] = gdf[col].apply(lambda x: x.strftime('%Y-%m-%d') if not pd.isna(x) else None)
                except:
                    gdf[col] = gdf[col].astype(str)

    return gdf


def extract_shp_from_zip(zip_path):
    """解压zip文件，返回第一个shp文件路径和cpg存在状态"""
    # 清理并创建临时目录
    safe_remove_temp_dir()
    os.makedirs(TEMP_DIR, exist_ok=True)

    # 解压ZIP（处理中文文件名）
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        for file_info in zip_ref.infolist():
            # 尝试修复中文文件名编码问题
            try:
                # 尝试用GBK解码（Windows常用）
                file_info.filename = file_info.filename.encode('cp437').decode('gbk')
            except:
                try:
                    # 尝试用UTF-8解码
                    file_info.filename = file_info.filename.encode('cp437').decode('utf-8')
                except:
                    pass
            zip_ref.extract(file_info, TEMP_DIR)

    # 递归查找所有.shp文件（处理多层目录）
    shp_files = glob.glob(os.path.join(TEMP_DIR, "**", "*.shp"), recursive=True)

    if not shp_files:
        raise ValueError("未找到.shp文件")

    # 检查.cpg文件是否存在（不区分大小写）
    shp_dir = os.path.dirname(shp_files[0])
    cpg_files = [f for f in os.listdir(shp_dir) if f.lower().endswith('.cpg')]
    cpg_exists = len(cpg_files) > 0

    return shp_files[0], cpg_exists  # 只返回第一个找到的shp文件


def is_projection_crs(prj_file_path):
    """检查坐标系类型，增强检测能力"""
    try:
        with open(prj_file_path, 'r', encoding='utf-8', errors='ignore') as prj_file:
            prj_content = prj_file.read().upper()

            # 检查是否明确标识为投影坐标系
            if "PROJCS" in prj_content:
                coordinate_type = "Projection"
                coordinate_system = prj_content.split('PROJCS["')[1].split('"')[
                    0] if 'PROJCS["' in prj_content else "Unknown"
                return coordinate_type, coordinate_system

            # 检查是否明确标识为地理坐标系
            if "GEOGCS" in prj_content:
                coordinate_type = "Geographic"
                coordinate_system = prj_content.split('GEOGCS["')[1].split('"')[
                    0] if 'GEOGCS["' in prj_content else "Unknown"
                return coordinate_type, coordinate_system

            # 尝试通过坐标单位判断
            if "UNIT[" in prj_content:
                unit_info = prj_content.split("UNIT[")[1].split("]")[0]
                if "DEGREE" in unit_info or "DEG" in unit_info:
                    return "Geographic", "Unknown Geographic System"
                if "METER" in unit_info or "METRE" in unit_info:
                    return "Projection", "Unknown Projected System"

            # 默认返回未知
            return "Unknown", "Unknown Coordinate System"
    except Exception as e:
        return "Error", f"读取.prj文件出错: {str(e)}"


def style_function(feature):
    """地块边界的显示样式"""
    return {
        'color': 'blue',
        'fillColor': 'transparent',
        'weight': 4
    }


def show_dataframe_in_window(df, window_title="边界文件检查结果"):
    """在Tkinter窗口中显示DataFrame"""
    root = tk.Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f"{screen_width // 2}x{screen_height // 2}")
    root.title(window_title)

    # 创建Treeview
    columns = list(df.columns)
    tv = ttk.Treeview(root, show="headings", columns=columns)

    # 设置列标题和宽度
    for col in columns:
        tv.heading(col, text=col, anchor='center')
        tv.column(col, width=100, anchor='center', stretch=True)

    # 添加滚动条
    scrollbar_y = ttk.Scrollbar(root, orient="vertical", command=tv.yview)
    scrollbar_x = ttk.Scrollbar(root, orient="horizontal", command=tv.xview)
    tv.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

    tv.pack(expand=True, fill="both")
    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")

    # 插入数据
    for idx, row in df.iterrows():
        tv.insert("", "end", values=row.tolist())

    root.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()


def extract_land_block_code(filename):
    """从文件名中提取13位地块编码"""
    # 匹配13位连续数字（前后不能是数字）
    match = re.search(r'(?<!\d)\d{13}(?!\d)', filename)
    if match:
        return match.group(0)
    return None


def main():
    # 选择文件夹
    folder_path = select_folder_path()
    excel_file = os.path.join(folder_path, "地块信息.xlsx")

    # 初始化统计变量
    pass_count = 0
    not_found_count = 0
    cpg_missing_count = 0
    field_missing_count = 0
    geometry_error_count = 0
    crs_error_count = 0
    in_polygon_count = 0

    # 检查Excel文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误：地块信息.xlsx文件不存在于 {folder_path}")
        sys.exit()

    # 读取原始地块信息
    try:
        original_df = pd.read_excel(excel_file)
        # 确保地块编码列为字符串类型，并去除空格
        original_df['地块编码'] = original_df['地块编码'].astype(str).str.strip()
        total_land_blocks = len(original_df)
        # print(f"地块信息Excel中共有 {total_land_blocks} 个地块编码")
    except Exception as e:
        print(f"读取地块信息.xlsx失败: {e}")
        sys.exit()

    # 确保必要列存在
    required_columns = ['地块编码', '经度', '纬度']
    missing_cols = [col for col in required_columns if col not in original_df.columns]
    if missing_cols:
        print(f"错误：地块信息.xlsx缺少必要列: {', '.join(missing_cols)}")
        sys.exit()

    # 创建结果DataFrame
    result_columns = [
        'zip_file_name', 'shp_file_relative', '地块编码', '地块名称', '经度', '纬度',
        'cpg', 'polygon', 'field', 'field_content', 'crs', 'In_polygon',
        '经度new', '纬度new', 'result'
    ]
    result_df = pd.DataFrame(columns=result_columns)
    for col in result_columns:
        result_df[col] = result_df[col].astype(str)

    # 创建地图
    m = folium.Map(location=[23.1, 113.25], zoom_start=10, control_scale=True,
                   tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                   attr='Esri')

    # 收集所有zip文件
    zip_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.zip')]

    # 按地块编码分组ZIP文件
    zip_by_land_block = defaultdict(list)
    for zf in zip_files:
        land_block_code = extract_land_block_code(zf)
        if land_block_code:
            zip_by_land_block[land_block_code].append(zf)

    # 处理每个匹配的ZIP文件（每个ZIP独立处理）
    for land_block_code, zip_list in zip_by_land_block.items():
        # 检查地块编码是否在原始数据中
        matching_rows = original_df[original_df['地块编码'] == land_block_code]
        if matching_rows.empty:
            # 记录未匹配到地块信息的ZIP
            for zip_file in zip_list:
                result_dict = {
                    'zip_file_name': zip_file,
                    '地块编码': land_block_code,
                    'result': '地块信息中未找到该地块编码'
                }
                # 使用pd.concat替代append
                result_df = pd.concat([result_df, pd.DataFrame([result_dict])], ignore_index=True)
            continue

        # 获取地块信息（取第一行，假设地块编码唯一）
        row = matching_rows.iloc[0]

        # 处理该地块编码对应的所有ZIP文件
        for zip_file in zip_list:
            zip_path = os.path.join(folder_path, zip_file)
            # print(f"正在处理: {zip_file} (地块编码: {land_block_code})")

            result_dict = {
                'zip_file_name': zip_file,
                '地块编码': land_block_code,
                '地块名称': row.get('地块名称', ''),
                '经度': row['经度'],
                '纬度': row['纬度'],
                'result': []  # 改为列表，用于累积所有问题
            }

            # 确保临时目录干净
            safe_remove_temp_dir()
            os.makedirs(TEMP_DIR, exist_ok=True)

            try:
                # 解压zip并查找shp文件
                try:
                    shp_path, cpg_exists = extract_shp_from_zip(zip_path)
                    # 获取SHP文件在ZIP中的相对路径（修复中文乱码）
                    shp_relative = os.path.relpath(shp_path, TEMP_DIR)
                    # 尝试修复中文路径乱码
                    try:
                        shp_relative = shp_relative.encode('latin1').decode('gbk')
                    except:
                        try:
                            shp_relative = shp_relative.encode('latin1').decode('utf-8')
                        except:
                            pass
                    result_dict['shp_file_relative'] = shp_relative
                except Exception as e:
                    result_dict['result'].append(f"解压错误: {str(e)}")
                    print(f"处理 {zip_file} 时出错:\n{traceback.format_exc()}")
                    # 使用pd.concat替代append
                    result_dict['result'] = " | ".join(result_dict['result']) if result_dict['result'] else "pass"
                    result_df = pd.concat([result_df, pd.DataFrame([result_dict])], ignore_index=True)
                    continue

                # 检查cpg文件（不区分大小写）
                if not cpg_exists:
                    result_dict['cpg'] = '缺失cpg文件'
                    result_dict['result'].append('cpg文件缺失')
                else:
                    result_dict['cpg'] = '是'

                # 读取shp文件（自动检测编码）
                gdf = None
                sf = None
                try:
                    # 检测编码
                    encoding = detect_encoding(shp_path)
                    # print(f"检测到SHP编码: {encoding} (地块编码: {land_block_code})")

                    # 读取shp
                    # gdf = gpd.read_file(shp_path, encoding=encoding)
                    try:
                        # 优先尝试指定编码
                        gdf = gpd.read_file(shp_path, encoding=encoding)
                    except:
                        # 如果失败，尝试使用errors='ignore'忽略编码错误
                        gdf = gpd.read_file(shp_path, encoding=encoding, errors='ignore')

                    # 显式关闭shapefile.Reader
                    sf = shapefile.Reader(shp_path, encoding=encoding)
                    # 立即读取编码并关闭
                    # encoding = sf.encoding
                    result_dict['encoding'] = sf.encoding
                    sf.close()
                    sf = None
                except Exception as e:
                    if sf:
                        try:
                            sf.close()
                        except:
                            pass
                    result_dict['result'].append(f"shp读取错误: {str(e)}")
                    # 使用pd.concat替代append
                    result_dict['result'] = " | ".join(result_dict['result']) if result_dict['result'] else "pass"
                    result_df = pd.concat([result_df, pd.DataFrame([result_dict])], ignore_index=True)
                    continue
                finally:
                    # 确保shapefile.Reader已关闭
                    if sf:
                        try:
                            sf.close()
                        except:
                            pass

                # 检查几何类型
                if gdf.empty or 'geometry' not in gdf.columns:
                    result_dict['result'].append('SHP文件无有效几何数据')
                else:
                    geom_types = gdf.geometry.geom_type.unique()
                    if 'Polygon' in geom_types or 'MultiPolygon' in geom_types:
                        result_dict['polygon'] = '是'
                    elif 'Point' in geom_types:
                        result_dict['polygon'] = 'point shp，请转为polygon shp'
                        result_dict['result'].append('几何类型错误')
                    else:
                        result_dict['polygon'] = 'line shp，请转为polygon shp'
                        result_dict['result'].append('几何类型错误')

                # 检查必要字段（支持中英文）
                required_fields_mapping = {
                    '地块名称': ['地块名称', 'DKMC', 'dkmc'],
                    '地块代码': ['地块代码', 'DKDM', 'DKBM', 'dkdm', 'dkbm'],
                    '行政区代码': ['行政区代码', 'XZQDM', 'xzqdm'],
                    '行政区名称': ['行政区名称', 'XZQMC', 'xzqmc'],
                    '地块面积': ['地块面积', 'YDMJ', 'ydmj']
                }

                found_fields = []
                for chinese_field, possible_names in required_fields_mapping.items():
                    for name in possible_names:
                        if name in gdf.columns:
                            found_fields.append(chinese_field)
                            break

                if len(found_fields) < 5:
                    missing = [field for field in required_fields_mapping.keys() if field not in found_fields]
                    result_dict['field'] = f'缺少字段：{", ".join(missing)}'
                    result_dict['result'].append('字段缺失')
                else:
                    result_dict['field'] = '是'

                # 检查字段内容
                empty_fields = []
                for chinese_field, possible_names in required_fields_mapping.items():
                    for name in possible_names:
                        if name in gdf.columns:
                            if gdf[name].isnull().all():
                                empty_fields.append(chinese_field)
                            break

                if empty_fields:
                    result_dict['field_content'] = f'字段内容为空：{", ".join(empty_fields)}'
                    result_dict['result'].append('字段内容为空')
                else:
                    result_dict['field_content'] = '是'

                # 检查坐标系
                prj_path = os.path.splitext(shp_path)[0] + ".prj"

                if os.path.exists(prj_path):
                    coord_type, coord_system = is_projection_crs(prj_path)
                    result_dict['crs'] = coord_system

                    # 地理坐标系警告（但不中断处理）
                    if coord_type == "Geographic":
                        result_dict['result'].append('地理坐标系（注意：应使用投影坐标系）')
                else:
                    result_dict['crs'] = '.prj文件不存在'
                    result_dict['result'].append('坐标系文件缺失')

                # 检查点是否在面内（无论坐标系类型）
                if 'geometry' in gdf.columns and not gdf.empty:
                    point = Point(row['经度'], row['纬度'])
                    point_series = gpd.GeoSeries([point], crs="EPSG:4490")  # 原始坐标系是CGCS2000地理坐标

                    # 转换点到shp的坐标系
                    try:
                        point_projected = point_series.to_crs(gdf.crs).iloc[0]
                    except Exception as e:
                        result_dict['result'].append(f'坐标系转换失败: {str(e)}')
                    else:
                        # 检查点是否在多边形内
                        within_polygon = False
                        for poly in gdf.geometry:
                            if point_projected.within(poly):
                                within_polygon = True
                                break

                        if within_polygon:
                            result_dict['In_polygon'] = '是'
                        else:
                            result_dict['In_polygon'] = '地块位置不在边界范围内'
                            result_dict['result'].append('地块位置不在边界范围内')

                            # 计算多边形中心
                            centroid = gdf.geometry.unary_union.centroid
                            try:
                                # 1. 首先检查.prj文件标识的坐标系类型
                                prj_path = os.path.splitext(shp_path)[0] + ".prj"
                                coord_type = "Unknown"
                                # coord_system = "Unknown"
                                if os.path.exists(prj_path):
                                    coord_type, _ = is_projection_crs(prj_path)

                                # 2. 检查实际坐标值范围来确认坐标系类型
                                # 如果坐标值在合理经纬度范围内，则认为是地理坐标
                                if -180 <= centroid.x <= 180 and -90 <= centroid.y <= 90:
                                    # is_geo = True
                                    coord_type = "Geographic"
                                    print(
                                        f"地块 {land_block_code}: 检测到实际为地理坐标系 (x={centroid.x:.6f}, y={centroid.y:.6f})")
                                else:
                                    # 如果坐标值超出经纬度范围，但.prj标识为地理坐标，则可能是错误
                                    if coord_type == "Geographic" and (abs(centroid.x) > 180 or abs(centroid.y) > 90):
                                        print(
                                            f"警告: 地块 {land_block_code} 的.prj标识为地理坐标系，但坐标值超出范围 (x={centroid.x:.6f}, y={centroid.y:.6f})")
                                        # 尝试将其视为投影坐标处理
                                        coord_type = "Projection"
                                    # is_geo = False

                                # 3. 根据实际坐标系类型处理
                                if coord_type == "Geographic":
                                    # 地理坐标系下，确保输出的是经纬度
                                    if -180 <= centroid.x <= 180 and -90 <= centroid.y <= 90:
                                        result_dict['经度new'] = round(centroid.x, 6)
                                        result_dict['纬度new'] = round(centroid.y, 6)
                                        print(
                                            f"地块 {land_block_code}: 使用地理坐标系值 (经度: {result_dict['经度new']}, 纬度: {result_dict['纬度new']})")
                                    else:
                                        # 虽然标识为地理坐标，但值超出范围，尝试转换
                                        print(f"地块 {land_block_code}: 地理坐标值异常，尝试转换")
                                        centroid_series = gpd.GeoSeries([centroid], crs=gdf.crs)
                                        centroid_cgcs = centroid_series.to_crs("EPSG:4490").iloc[0]
                                        result_dict['经度new'] = round(centroid_cgcs.x, 6)
                                        result_dict['纬度new'] = round(centroid_cgcs.y, 6)
                                        print(
                                            f"地块 {land_block_code}: 转换后坐标 (经度: {result_dict['经度new']}, 纬度: {result_dict['纬度new']})")
                                else:
                                    # 投影坐标系下，需要转换回地理坐标
                                    centroid_series = gpd.GeoSeries([centroid], crs=gdf.crs)
                                    centroid_cgcs = centroid_series.to_crs("EPSG:4490").iloc[0]
                                    result_dict['经度new'] = round(centroid_cgcs.x, 6)
                                    result_dict['纬度new'] = round(centroid_cgcs.y, 6)
                                    print(
                                        f"地块 {land_block_code}: 投影坐标转换 (经度: {result_dict['经度new']}, 纬度: {result_dict['纬度new']})")

                                # 4. 验证新坐标是否在合理范围内
                                if not (-180 <= result_dict['经度new'] <= 180) or not (
                                        -90 <= result_dict['纬度new'] <= 90):
                                    print(
                                        f"警告: 地块 {land_block_code} 的新坐标超出合理范围 (经度: {result_dict['经度new']}, 纬度: {result_dict['纬度new']})")
                                    result_dict['result'].append('新坐标计算异常')

                            except Exception as e:
                                result_dict['result'].append(f'中心点坐标计算失败')
                                print(f"计算地块 {land_block_code} 的中心点坐标时出错: {str(e)}")
                else:
                    result_dict['result'].append('无法检查点是否在多边形内（无有效几何数据）')

                # === 关键修复：转换Timestamp对象为字符串 ===
                gdf = convert_timestamps_to_strings(gdf)

                # 添加到地图（为每个ZIP创建独立图层）
                layer_name = f"{zip_file} ({land_block_code})"
                popup_content = folium.Popup(
                    f"<b>ZIP文件:</b> {zip_file}<br>"
                    f"<b>地块编码:</b> {land_block_code}<br>"
                    + pd.DataFrame(gdf.drop(columns='geometry', errors='ignore')).to_html(),
                    max_width=1200
                )
                folium.GeoJson(
                    gdf,
                    style_function=style_function,
                    tooltip=layer_name,
                    popup=popup_content,
                    name=layer_name
                ).add_to(m)

                # 添加点标记（只添加一次，避免重复）
                if zip_file == zip_list[0]:  # 只在第一个ZIP时添加
                    folium.Marker(
                        [row['纬度'], row['经度']],
                        popup=f"{land_block_code}<br>{row.get('地块名称', '')}",
                        icon=folium.Icon(color='lightblue', icon='info-sign'),
                        tooltip=land_block_code
                    ).add_to(m)

            except Exception as e:
                result_dict['result'].append(f"处理错误: {str(e)}")
                print(f"处理 {zip_file} 时出错:\n{traceback.format_exc()}")

            finally:
                # 确保临时目录被清理
                safe_remove_temp_dir()

            # 将所有问题汇总到result字段
            result_dict['result'] = " | ".join(result_dict['result']) if result_dict['result'] else "pass"
            result_df = pd.concat([result_df, pd.DataFrame([result_dict])], ignore_index=True)

    # 确保最终清理临时目录
    safe_remove_temp_dir()

    # 添加图层控制（如果地图上有多个图层）
    folium.LayerControl().add_to(m)

    # 保存结果到Excel的新sheet
    try:
        # 如果Excel文件已有"result"工作表，先删除它
        try:
            from openpyxl import load_workbook
            book = load_workbook(excel_file)
            if 'result' in book.sheetnames:
                del book['result']
                book.save(excel_file)
        except Exception as e:
            print(f"清理Excel工作表时出错: {str(e)}")

        # 写入新结果
        with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='result', index=False)
        # 获取pass的数量
        pass_count = result_df[result_df['result'] == 'pass'].shape[0]
        # 获取"地块信息中未找到该地块编码"的数量
        not_found_count = result_df[result_df['result'].str.contains('地块信息中未找到该地块编码', na=False)].shape[0]
        # 获取cpg文件缺失的数量
        cpg_missing_count = result_df[result_df['result'].str.contains('cpg文件缺失', na=False)].shape[0]
        # 获取字段缺失的数量
        field_missing_count = result_df[result_df['result'].str.contains('字段缺失|字段内容为空', na=False)].shape[
            0]
        # 获取几何类型错误的数量
        geometry_error_count = result_df[result_df['result'].str.contains('几何类型错误', na=False)].shape[0]
        # 获取坐标系问题的数量
        crs_error_count = result_df[result_df['result'].str.contains('坐标系文件缺失|地理坐标系', na=False)].shape[
            0]
        in_polygon_count = result_df[result_df['result'].str.contains('地块位置不在边界范围内', na=False)].shape[0]
        # 创建统计信息
        stats_data = {
            '统计项': [
                '地块信息的地块编码数量',
                'zip文件数量',
                '去重地块编码后zip文件数量',
                'PASS地块数量',
                '未对应地块编码的zip数量',
                'cpg文件缺失的数量',
                '字段问题的数量',
                '几何类型错误的数量',
                '坐标系问题的数量',
                '地块位置不在边界范围的数量'
            ],
            '数量': [
                total_land_blocks,
                len(zip_files),
                len(zip_by_land_block),
                pass_count,
                not_found_count,
                cpg_missing_count,
                field_missing_count,
                geometry_error_count,
                crs_error_count,
                in_polygon_count
            ]
        }
        stats_df = pd.DataFrame(stats_data)
        # 打印统计信息
        print("\n" + "=" * 60)
        print("边界文件检查情况")
        print("=" * 60)
        print(f"{'统计项':} | {'数量'}")
        print("-" * 45)

        # 格式化输出每个统计项
        print(f"{'地块信息的地块编码数量':} | {total_land_blocks:}")
        print(f"{'zip文件数量':} | {len(zip_files):}")
        print(f"{'去重地块编码后zip文件数量':} | {len(zip_by_land_block):}")
        print(f"{'PASS地块数量':} | {pass_count:}")
        print(f"{'未对应地块编码的zip数量':} | {not_found_count:}")
        print(f"{'cpg文件缺失的数量':} | {cpg_missing_count:>}")
        print(f"{'字段问题的数量':} | {field_missing_count:>}")
        print(f"{'几何类型错误的数量':} | {geometry_error_count:}")
        print(f"{'坐标系问题的数量':} | {crs_error_count:}")
        print(f"{'地块位置不在边界范围的数量':} | {in_polygon_count:}")
        print("-" * 45)
        # 将统计信息保存到新的工作表
        with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            stats_df.to_excel(writer, sheet_name='统计信息', index=False)
        print(f"检查结果已保存到 {excel_file} 的 'result' 工作表")
        # print(len(result_df[(result_df['shp_file_relative'].fillna('').astype(str).str.strip() == '') & (
        #             result_df['地块编码'].fillna('').astype(str).str.strip() != '')]))
        print(f"统计信息已保存到 {excel_file} 的 '统计信息' 工作表")

    except Exception as e:
        print(f"保存结果时出错: {str(e)}")

    # 显示结果
    show_dataframe_in_window(result_df, "边界文件检查结果")

    # 保存地图
    map_path = os.path.join(folder_path, "地块边界检查结果.html")
    try:
        m.save(map_path)
        print(f"地图已保存至: {map_path}")

        # 验证HTML文件是否为空
        if os.path.exists(map_path) and os.path.getsize(map_path) < 1000:
            print("警告：生成的HTML地图文件可能为空，请检查")
            # 创建一个简单的HTML文件作为备份
            backup_html = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>地块边界检查结果</title>
                <meta charset="utf-8" />
            </head>
            <body>
                <h1>地块边界检查结果</h1>
                <p>地图生成失败，请检查日志。</p>
                <p>可能原因：</p>
                <ul>
                    <li>没有有效的SHP文件</li>
                    <li>坐标系设置不当</li>
                    <li>网络问题（无法加载底图）</li>
                </ul>
            </body>
            </html>
            """
            with open(map_path, 'w', encoding='utf-8') as f:
                f.write(backup_html)
    except Exception as e:
        print(f"保存地图时出错: {str(e)}")
        # 创建一个简单的HTML文件作为备份
        backup_html = os.path.join(folder_path, "地块边界检查结果.html")
        with open(backup_html, 'w', encoding='utf-8') as f:
            f.write("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>地块边界检查结果</title>
                <meta charset="utf-8" />
            </head>
            <body>
                <h1>地块边界检查结果</h1>
                <p>地图生成失败，请检查日志。</p>
                <p>可能原因：</p>
                <ul>
                    <li>没有有效的SHP文件</li>
                    <li>坐标系设置不当</li>
                    <li>网络问题（无法加载底图）</li>
                </ul>
            </body>
            </html>
            """)
        print(f"已创建备用HTML文件: {backup_html}")


if __name__ == "__main__":
    main()
